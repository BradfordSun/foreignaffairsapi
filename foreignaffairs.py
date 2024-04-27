import datetime
import os

from fastapi import HTTPException
import pymysql
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import StreamingResponse
import asyncio
import json

from fastapi.middleware.cors import CORSMiddleware
from openpyxl.reader.excel import load_workbook

app = FastAPI()
# 跨域相关设置
origins = [
    "*"
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db_connection():
    connection = pymysql.connect(host=os.getenv("dbhost"),
                                 user=os.getenv("dbuser"),
                                 password=os.getenv("dbpassword"),
                                 database=os.getenv("dbname"),
                                 charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)
    return connection

def timecompare(timelist):
    # 如果列表为空，表示没有护照信息
    try:
        if not timelist:
            return "无护照信息"

        # 初始化最早的日期为很旧的一个日期
        newest_time = datetime.datetime.strptime("1900-01-01", "%Y-%m-%d")

        # 遍历日期列表，找出最新的日期
        for t in timelist:
            compare_time = datetime.datetime.strptime(t, "%Y-%m-%d")
            if compare_time > newest_time:
                newest_time = compare_time

        # 获取当前日期，并格式化
        now = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        # 如果最新日期早于当前日期，表示护照已过期
        if newest_time < now:
            return "护照均已过期"

        # 计算最新日期与当前日期的差距
        difference = newest_time - now
    except Exception as e:
        print(e)

    # 如果护照离过期不到6个月，就提示即将到期
    if difference < datetime.timedelta(days=6 * 30):  # 近似每月30天
        return newest_time.strftime("%Y-%m-%d") + "(即将到期)"
    else:
        # 否则，返回护照的最新到期日期
        return newest_time.strftime("%Y-%m-%d")

# 获取文件更新时间API
@app.get("/fileupdate")
def file_update():
    connection = None
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            # 数据库只有一行数据
            cursor.execute("SELECT passportupdate,workupdate,basicupdate,baseupdate FROM fileupdate WHERE id = 1;")
            update_result = cursor.fetchone()
            passportupdate = update_result["passportupdate"]
            workupdate = update_result["workupdate"]
            basicupdate = update_result["basicupdate"]
            baseupdate = update_result["baseupdate"]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if connection:
            connection.close()
    return {"passportupdate":passportupdate, "workupdate":workupdate, "basicupdate":basicupdate, "baseupdate":baseupdate}

@app.get("/expire")
def expire_date():
    expire_items = []
    connection = None
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            # 获取每个人的最晚出国时间和最晚回国时间
            cursor.execute("""
                SELECT name,
                       MAX(departure_date) AS max_departure_date,
                       MAX(return_date) AS max_return_date
                FROM base
                GROUP BY name
            """)
            people = cursor.fetchall()

            # 检查每个人是否仍在海外
            for person in people:
                # 设置默认时间为1900-01-01，以免后面报错
                default_date = datetime.datetime.strptime('1900-01-01', '%Y-%m-%d').date()
                # 定义最晚的出国和回国日期
                max_departure = person['max_departure_date'] if person[
                                                                    'max_departure_date'] is not None else default_date
                max_return = person['max_return_date'] if person['max_return_date'] is not None else default_date

                if max_departure > max_return:
                    # 获取这个人入境时间最晚的记录的国家
                    cursor.execute("""
                        SELECT country
                        FROM base
                        WHERE name = %s AND entry_date = (
                            SELECT MAX(entry_date)
                            FROM base
                            WHERE name = %s
                        )
                    """, (person['name'], person['name']))
                    latest_country = cursor.fetchone()
                    if latest_country:
                        country = latest_country['country']
                        # 查询这个人的批件情况
                        cursor.execute("""
                                                    SELECT MAX(back_date) AS last_back_date
                                                    FROM external
                                                    WHERE name = %s AND country = %s
                                                """, (person['name'], country))
                        permission = cursor.fetchone()
                        today = datetime.datetime.now().date()
                        if permission and permission['last_back_date']:
                            # 查询这个人的部门情况
                            cursor.execute("""
                                                    SELECT department
                                                    FROM basic
                                                    WHERE name = %s
                                                """, (person['name'],))
                            department_info = cursor.fetchone()
                            department = department_info['department'] if department_info else '未知部门'
                            # 找到这个国家的最后批件日期
                            last_back_date = permission['last_back_date']
                            if last_back_date < today:
                                status = f"{last_back_date} (已过期)"
                            elif (last_back_date - today).days < 60:
                                status = f"{last_back_date} (即将过期)"
                            else:
                                # 批件有效期内且离过期时间大于60天就不管
                                continue
                        else:
                            status = "无批件"
                        expire_items.append({
                            'Name': person['name'],
                            'Department':department,
                            'Country': country,
                            'PermissionStatus': status
                        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if connection:
            connection.close()
    return {"expire":expire_items}


# 获取护照台账和批件信息API
@app.get("/personal")
def personal_infos(name: str):
    connection = None
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            # 查询passport表中的数据
            cursor.execute("SELECT * FROM passport WHERE name = %s", (name,))
            passport_result = cursor.fetchall()

            passports = []
            for p in passport_result:
                expire_date = p['expire_date'].strftime('%Y-%m-%d')
                if datetime.datetime.strptime(expire_date, '%Y-%m-%d') < datetime.datetime.now():
                    expire_date += " (已过期)"
                else:
                    expire_date += " (有效)"

                passports.append({
                    "Name": p['name'],
                    "Gender": p['gender'],
                    "Birthday": p['birth_day'].strftime('%Y-%m-%d'),
                    "Birthplace": p['birth_place'],
                    "PassportNumber": p['passport_no'],
                    "PassportIssue": p['issue_date'].strftime('%Y-%m-%d'),
                    "PassportExpire": expire_date
                })

            # 查询external表中的数据
            cursor.execute("SELECT * FROM external WHERE name = %s", (name,))
            external_result = cursor.fetchall()

            infos = []
            for e in external_result:
                back_date = e['back_date'].strftime('%Y-%m-%d')
                if datetime.datetime.strptime(back_date, '%Y-%m-%d') < datetime.datetime.now():
                    back_date += " (失效)"
                else:
                    back_date += " (有效)"

                infos.append({
                    "Country": e['country'],
                    "ApprovalStart": e['leave_date'].strftime('%Y-%m-%d'),
                    "ApprovalEnd": back_date,
                    "ApproveNumber": e['approve_number'],
                })

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if connection:
            connection.close()

    return {"passport": passports, "info": infos}

# 获取部门团组API
@app.get("/department")
def department_infos(department: str):
    department_items = []
    connection = None
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            # 获取部门下所有人的信息
            cursor.execute("SELECT name FROM basic WHERE department = %s", (department,))
            departments = cursor.fetchall()
            names = [d['name'] for d in departments]
            for name in names:
                visa_information = ""
                cursor.execute("SELECT country, back_date FROM external WHERE name = %s", (name,))
                for d in cursor.fetchall():
                    back_date = d['back_date'].strftime('%Y-%m-%d')
                    if datetime.datetime.strptime(back_date, '%Y-%m-%d') >= datetime.datetime.now():
                        visa_information += f"{d['country']}: {back_date}|"

                # 查询护照信息
                cursor.execute("SELECT expire_date FROM passport WHERE name = %s", (name,))
                passport_dates = [p['expire_date'].strftime('%Y-%m-%d') for p in cursor.fetchall()]
                print(passport_dates)
                passport_date = timecompare(passport_dates)

                department_items.append(
                    {"Name": name, "Passportexpire": passport_date, "Approval": visa_information})

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if connection:
            connection.close()

    return department_items


# 以下是上传文件：
# 存储最新进度信息
progress_data = {"progress": 0, "step": "", "status": "", "resultText": ""}
update_event = asyncio.Event()


async def update_progress_data(progress, step, status, resultText=""):
    global progress_data
    progress_data = {
        "progress": progress,
        "step": step,
        "status": status,
        "resultText": resultText
    }
    update_event.set()

# 上传文件API
@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    filename = file.filename
    temp_file_path = f"./temp_{filename}"
    try:
        # 保存文件
        await update_progress_data(0, "保存文件", '', "")
        await asyncio.sleep(2)
        with open(temp_file_path, 'wb') as out_file:
            # Read the content of the uploaded file asynchronously
            while content := await file.read(1024 * 1024):  # Reading in chunks of 1MB
                out_file.write(content)  # Write the content to local file
    except Exception as e:
        await update_progress_data(100, "", 'error', f"保存文件失败，报错详情为：{str(e)}")
    await update_progress_data(20, "解析文件", '', "")
    await asyncio.sleep(1)
    # 四种文件分别调用解析和入库的函数，中间加个sleep让进度条更明显
    if filename == "国际公司因公护照信息-sun.xlsx":
        passport_data = await parse_passport_excel(temp_file_path)
        await update_progress_data(60, "数据入库", '', "")
        await asyncio.sleep(1)
        await store_passport_data(passport_data)
        await update_progress_data(100, "", 'success', "处理成功，所有数据已入库。")
    if filename == "基础信息维护-sun.xlsx":
        basic_data = await parse_basic_excel(temp_file_path)
        await update_progress_data(60, "数据入库", '', "")
        await asyncio.sleep(1)
        await store_basic_data(basic_data)
        await update_progress_data(100, "", 'success', "处理成功，所有数据已入库。")
    if filename == "国际公司外事工作台账-sun.xlsx":
        work_data = await parse_work_excel(temp_file_path)
        await update_progress_data(60, "数据入库", '', "")
        await asyncio.sleep(1)
        await store_work_data(work_data)
        await update_progress_data(100, "", 'success', "处理成功，所有数据已入库。")
    if filename == "常驻.xlsx":
        base_data = await parse_base_excel(temp_file_path)
        await update_progress_data(60, "数据入库", '', "")
        await asyncio.sleep(1)
        await store_base_data(base_data)
        await update_progress_data(100, "", 'success', "处理成功，所有数据已入库。")


@app.get("/events")
async def events():
    async def event_generator():
        while True:
            await update_event.wait()
            yield f"data: {json.dumps(progress_data)}\n\n"
            update_event.clear()
            if progress_data['status'] in ['success', 'error']:
                break

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# 处理国际公司因公护照信息-sun.xlsx
async def parse_passport_excel(file_path):
    passport_data = []
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook['all-in-one']
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if row[0] is None:
                # 姓名为0等于这行不存在
                continue

            name, gender, birthday, birthplace, passport_no, issue_date, expire_date, status, comment1, comment2, borrow_date = (
                row[0], row[1], f"{str(row[2])[:4]}-{str(row[2])[4:6]}-{str(row[2])[6:]}", row[3],
                row[4], f"{str(row[5])[:4]}-{str(row[5])[4:6]}-{str(row[5])[6:]}",
                f"{str(row[6])[:4]}-{str(row[6])[4:6]}-{str(row[6])[6:]}", row[8], row[9], row[10], row[11]
            )
            if status is None:
                status = ""
            if comment1 is None:
                comment1 = ""
            if comment2 is None:
                comment2 = ""
            if borrow_date is None:
                borrow_date = "1900-01-01"
            passport_info = {"name": name, "gender": gender, "birthday": birthday, "birthplace": birthplace,
                             "passport_no": passport_no, "issue_date": issue_date, "expire_date": expire_date,
                             "status": status, "comment1": comment1, "comment2": comment2, "borrow_date": borrow_date}
            passport_data.append(passport_info)
    except Exception as e:
        await update_progress_data(100, "", 'error', f"解析文件失败，报错详情为：{str(e)}")
    await asyncio.sleep(1)
    return passport_data


async def store_passport_data(passport_data):
    connection = None
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM passport")
            for p in passport_data:
                cursor.execute(
                    "INSERT INTO passport (name, gender, birth_day, birth_place, passport_no, issue_date, expire_date, status, comment1, comment2, borrow_date) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (p["name"], p["gender"], p["birthday"], p["birthplace"],
                     p["passport_no"], p["issue_date"], p["expire_date"],
                     p["status"], p["comment1"], p["comment2"],
                     p["borrow_date"]))
            cursor.execute("UPDATE fileupdate SET passportupdate = NOW() WHERE id = 1;")
            connection.commit()
    except Exception as e:
        await update_progress_data(100, "", 'error', f"数据入库失败，报错详情为：{str(e)}")
        if connection:
            connection.rollback()
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
    await asyncio.sleep(2)


async def parse_basic_excel(file_path):
    basic_data = []
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook['Sheet1']
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if row[1] is None:
                # 姓名为空等于这行不存在
                continue
            name = row[1]
            department = row[2]
            basic_info = {"name": name, "department": department}
            basic_data.append(basic_info)
    except Exception as e:
        await update_progress_data(100, "", 'error', f"解析文件失败，报错详情为：{str(e)}")
    await asyncio.sleep(1)
    return basic_data


async def store_basic_data(basic_data):
    connection = None
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM basic")
            for b in basic_data:
                cursor.execute(
                    "INSERT INTO basic (name, department) "
                    "VALUES (%s, %s)",
                    (b["name"], b["department"]))
            cursor.execute("UPDATE fileupdate SET basicupdate = NOW() WHERE id = 1;")
            connection.commit()
    except Exception as e:
        await update_progress_data(100, "", 'error', f"数据入库失败，报错详情为：{str(e)}")
        if connection:
            connection.rollback()
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
    await asyncio.sleep(2)


async def parse_work_excel(file_path):
    work_data = []
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook['sheet1']
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if row[3] is None or row[4] is None:
                # 没有国家或者没有团队成员说明这行没意义
                continue
            countries = split_values(row[3])
            members = split_values(row[4])
            for country in countries:
                for member in members:
                    leaveDate = format_date(row[6])
                    backDate = format_date(row[7])
                    applyType = row[8]
                    approveNumber = row[12]
                    work_info = {
                        "name": member,
                        "country": country,
                        "leaveDate": leaveDate,
                        "backDate": backDate,
                        "applyType": applyType,
                        "approveNumber": approveNumber
                    }
                    work_data.append(work_info)
    except Exception as e:
        await update_progress_data(100, "", 'error', f"解析文件失败，报错详情为：{str(e)}")
    await asyncio.sleep(1)
    return work_data


def split_values(raw_value):
    """ Split the raw value by the delimiter if it contains one, otherwise return it as a single-element list. """
    return raw_value.split('、') if '、' in raw_value else [raw_value]


def format_date(date_value):
    """ Return a formatted date if the input is not empty or a dash, otherwise return the default date. """
    return date_value if date_value not in ["", "-"] else "1900-01-01"


async def store_work_data(work_data):
    connection = None
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM external")
            for w in work_data:
                cursor.execute(
                    "INSERT INTO external (name, country, leave_date, back_date, apply_type, approve_number) "
                    "VALUES (%s, %s, %s, %s, %s, %s)",
                    (w["name"], w["country"], w["leaveDate"], w["backDate"], w["applyType"], w["approveNumber"]))
            cursor.execute("UPDATE fileupdate SET workupdate = NOW() WHERE id = 1;")
            connection.commit()
    except Exception as e:
        await update_progress_data(100, "", 'error', f"数据入库失败，报错详情为：{str(e)}")
        if connection:
            connection.rollback()
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
    await asyncio.sleep(2)

async def parse_base_excel(file_path):
    base_data = []
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook["常驻"]

        # 从第三行开始
        for i, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if row[2] is None:
                # 这里第三列是姓名，和其他的不一样。如果姓名为空，就抛异常
                raise ValueError(f"第{i}行的数据不完整，姓名缺失。")
            name = row[2]
            country = row[4]
            departure_date = row[5]
            return_date = row[6]
            entry_date = row[7]
            exit_date = row[8]
            base_info = {"name": name, "country":country, "departure_date":departure_date, "return_date":return_date, "entry_date":entry_date, "exit_date":exit_date}
            base_data.append(base_info)
    except Exception as e:
        await update_progress_data(100, "", 'error', f"解析文件失败，报错详情为：{str(e)}")
    await asyncio.sleep(1)
    return base_data


async def store_base_data(base_data):
    connection = None
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            cursor.execute(f"DELETE FROM base")
            for b in base_data:
                cursor.execute("INSERT INTO base (name,country,departure_date,return_date,entry_date,exit_date)" 
                               "VALUES (%s, %s, %s, %s, %s, %s)",
                               (b["name"], b["country"], b["departure_date"], b["return_date"], b["entry_date"], b["exit_date"]))
            cursor.execute("UPDATE fileupdate SET baseupdate = NOW() WHERE id = 1;")
            connection.commit()
    except Exception as e:
        await update_progress_data(100, "", 'error', f"数据入库失败，报错详情为：{str(e)}")
        if connection:
            connection.rollback()
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
    await asyncio.sleep(2)
